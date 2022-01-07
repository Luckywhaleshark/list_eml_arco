from re import X
import cv2
import os
from PIL import Image
import pytesseract
import openpyxl
file_list = []
file_list2 = []
file_list3 = []
a = 0

path = r'C:\py\test1' # 원본 jpg 파일 저장위치
path_save = r'C:\py\test2' # 전처리한 jpg 파일 저장위치
path_txt = r'C:\py\test3' # jpg -> txt 으로 oxr 한 txt 저장위치
file_len = os.listdir(path) # 원본 파일 개수

wb = openpyxl.Workbook() # 엑셀 새파일 만들기
wb.save('test.xlsx')    #엑셀 파일 저장


def ocr(file_list) :  # tesseract ocr 작동 함수 (file_list는 ocr 할 파일)
    file_txt = path_txt + '\\'  + file_list[:-3] + 'txt' # ocr한후 txt 파일 저장 이름
    file_path1 = path_save + '\\' + file_list # ocr할 파일 이름
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    text = pytesseract.image_to_string(Image.open(file_path1)) 
    with open(file_txt,"w",encoding="utf8") as file : #ocr이후 txt 저장할 파일 열고 저장
        file.write(text)

for x in os.listdir(path) : #path 내의 jpg 파일 수 
    if x.endswith('jpg') :
        file_list.append(x)

print(file_list)


for x in file_list :            # jpg 파일 전처리 
    file_path = path + '\\' + x
    file_save_path = path_save + '\\' + x
    src = cv2.imread(file_path)
    dst1 = cv2.inRange(src, (250,250,250), (255,255,255))
    cv2.imwrite(file_save_path,dst1)

for x in os.listdir(path_save) :
    if x.endswith('jpg') :
        file_list2.append(x)

for x in file_list2 :
    file_save_path = path_save + '\\' + x
    img = Image.open(file_save_path)
    img_size = img.size
    img_size_left = img_size[0]
    img_size_top = img_size[1]
    img_area = (img_size_left*0.25,1,img_size_left*0.75,img_size_top * 0.45)
    cropped_img = img.crop(img_area)
    cropped_img.save(file_save_path)


for x in file_list2 :
     ocr(x)

for x in os.listdir(path_txt) :
    if x.endswith('txt') :
        file_list3.append(x)

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.active
for x in file_list3 :
    file_txt_path = path_txt + '\\' + x
    a = a + 1
    print(file_txt_path)
    with open(file_txt_path,'r',encoding='utf8') as f :
        line = f.readline() 
        sheet.cell(row=a,column=1).value = x[:-4]
        sheet.cell(row=a,column=2).value = line
    
wb.save('test.xlsx')