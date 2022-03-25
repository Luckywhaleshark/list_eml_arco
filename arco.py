# -*- coding: euc-kr -*-
from re import X
from PIL import Image, ImageChops
from datetime import  datetime
import requests
import time
import pyautogui
import clipboard
import os
import cv2
import subprocess
import openpyxl
import pytesseract
import numpy as np
import sys

num = 0         #기본적인 변수들 선언
num1 = 0
a = 0
b = 0
uniq = 1
file_list = []
copy_result = None
copy_result_url = None
directory = ['jpg','jpg_crop','error','txt']
today = datetime.now()
today_date = today.strftime("%y.%m.%d")
#today_date = "22.03.21"
print(today_date)
today_path = r'C:\py\venv'+ '\\' + today_date


def download(url, file_name=None):
    if not file_name:
        file_name = url.split('/')[-1]
    r = requests.get(url)
    file_path = adr_jpg + '\\' + file_name
    file = open(file_path, "wb")
    file.write(r.content)
    file.close()
    time.sleep(0.5)

def copy_jpg() : # url 추출하는 함수
    time.sleep(0.1)
    pyautogui.click(x=-1295, y=371)
    pyautogui.hotkey('ctrl', 'u')
    time.sleep(0.1)
    pyautogui.click(x=-1295, y=371)
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.1)

def open_file_eml(address) : # address의 파일을 실행
    subprocess.run(address, shell=True)

def file_list_ext(address,ext) : # 파일 (address)에서 eml확장자 파일 file_list에 리스트 추가
    file_list.clear()
    file = os.listdir(address)
    for x in file :
        if x.endswith(ext):
            file_list.append(x)

def kor_split (str) :
    num2 = len(str)
    kor_name = ''
    for i in range(0,num2) :
        str_to_asc = ord(str[i])
        if str_to_asc >= 44032 :
            kor_name = kor_name + str[i]
    return kor_name

def ocr(file_list,input_path,output_path) :  # tesseract ocr 작동 함수 (file_list는 ocr 할 파일)
    file_txt = output_path + '\\'  + file_list[:-3] + 'txt' # ocr한후 txt 파일 저장 이름
    file_path1 = input_path + '\\' + file_list # ocr할 파일 이름
    config = '-l kor+eng --oem 3 --psm 11'
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    text = pytesseract.image_to_string(Image.open(file_path1),config=config)
    with open(file_txt,"w",encoding="utf8") as file : #ocr이후 txt 저장할 파일 열고 저장
        file.write(text)
    print(file_txt)

def image_preprocessing(file_list) :
    for x in file_list :
        file_path = adr_jpg + '\\' + x
        file_save_path = adr_jpg_crop + '\\' + x
        print(file_path)
        print(file_save_path)
        src = cv2.imread(file_path)
        dst1 = cv2.inRange(src, (240,240,240), (255,255,255))   # 흰부분 제외한 부분은 검은색으로 처리
        cv2.imwrite(file_save_path,dst1)                        # 검은색으로 처리한 부분 저장
        img = Image.open(file_save_path)
        img_size = img.size
        img_size_left = img_size[0]
        img_size_top = img_size[1]
        img_area = (img_size_left * 0.25, 1, img_size_left * 0.75, img_size_top * 0.45)
        cropped_img = img.crop(img_area)
        cropped_img.save(file_save_path)


def txt_to_excel(file_list) :
    wb = openpyxl.load_workbook('test.xlsx')
    sheet = wb.active
    for x in file_list:
        file_txt_path = adr_txt + '\\' + x
        a = a + 1
        print(file_txt_path)
        with open(file_txt_path, 'r', encoding='utf8') as f:
            line = f.readline()
            sheet.cell(row=a, column=3).value = x[:-4]
            sheet.cell(row=a, column=4).value = line
    wb.save('test.xlsx')

def png_to_jpg(file_list) :
    for x in file_list :
        file_png_path = adr_jpg + '\\' + x
        file_jpg_path = adr_jpg + '\\' + x[:-3] + 'jpg'
        im = Image.open(file_png_path).convert('RGB')
        im.save(file_jpg_path, 'jpeg')
        os.remove(file_png_path)

def erosion_dilation(file_list,input_path,output_path): #확장자 무시하고 파일명만 따서 하는 함수, 확장자 상관없음
    filename_input = input_path + '\\' + file_list[:-3] + 'jpg'
    filename_output = output_path + '\\' + file_list[:-3] + 'jpg'
    image = cv2.imread(filename_input, cv2.IMREAD_GRAYSCALE)  #// 회색조로 이미지 객체를 생서한다.

    #// make kernel matrix for dilation and erosion (Use Numpy)
    kernel_size_row = 3
    kernel_size_col = 3
    kernel = np.ones((3, 3), np.uint8)

    erosion_image = cv2.erode(image, kernel, iterations=1)  #// make erosion image
    cv2.imwrite(filename_output,erosion_image)


if os.path.isdir(today_path) == False :
    print('폴더가 없습니다. 생성합니다.')
    os.mkdir(today_path)
    sys.exit()
else :
    directory_eml_path = today_path + '\\' + 'eml_' + today_date
    if os.path.isdir(directory_eml_path) == True :
        file_list_ext(directory_eml_path, 'eml')
        if len(file_list) > 0 :
            for x in directory :
                directory_path = today_path + '\\' + x + '_' + today_date
                if os.path.isdir(directory_path) == False :
                        os.mkdir(directory_path)
                        print('{} 가 생성되었습니다.'.format(directory_path))
        else :
            print("eml파일을 eml_{} 폴더에 넣으세요".format(today_date))
            sys.exit()
    else :
        print("eml_{} 폴더가 없습니다. eml 파일을 생성된 폴더에 넣으세요.".format(today_date))
        os.mkdir(directory_eml_path)
        sys.exit()

adr_eml = today_path + '\\' + 'eml_' + today_date # eml 파일 주소
adr_jpg = today_path + '\\' + 'jpg_' + today_date # jpg 저장파일주소
adr_jpg_crop = today_path + '\\' + 'jpg_crop_' + today_date # jpg_crop 저장파일주소
adr_ocr_error = today_path + '\\' + 'error_' + today_date#오류 데이터 저장위치
adr_txt = today_path + '\\' + 'txt_' + today_date # txt 저장주소

wb = openpyxl.Workbook() # 엑셀 새파일 만들기
wb.save('arco_{}.xlsx'.format(today_date))    #엑셀 파일 저장

wb = openpyxl.load_workbook('arco_{}.xlsx'.format(today_date))
sheet = wb.active

file_list_ext(adr_eml,'eml')
for file_eml in file_list :                     # eml에서 jpg,png파일 다운로드
    file_address_eml = adr_eml + '\\' + file_eml
    b=b+1
    open_file_eml(file_address_eml)
    copy_jpg()
    copy_result = clipboard.paste()
    copy_result_url = copy_result.split('\'') # ' 작은따음표 기준으로 나눔
    file_name = copy_result_url[1].split('/')[-1]
    name_split = copy_result_url[0]
    name = kor_split(name_split)
    file_path = adr_jpg + '\\' + file_name
    while os.path.exists(file_path) :
        file_name = file_name[0:-4] + '(%d)'%uniq + file_name[-4:]
        uniq +=1
        break
    print(name)
    print(file_name)
    download(copy_result_url[1],file_name)
    pyautogui.hotkey('ctrl', 'w')
    pyautogui.hotkey('ctrl', 'w')
    sheet.cell(row=b, column=1).value = name
    sheet.cell(row=b, column=2).value = file_name
    copy_result_url.clear()

wb.save('arco_{}.xlsx'.format(today_date))    # jpg,png 추출 완료



file_list_ext(adr_jpg,'png')
png_to_jpg(file_list)           #png 파일 jpg로 변환

file_list_ext(adr_jpg,'jpg')
image_preprocessing(file_list)


file_list_ext(adr_jpg_crop,'jpg') # crop 한 jpg를 txt로 ocr
for x in file_list :
    ocr(x,adr_jpg_crop,adr_txt)

file_list_ext(adr_txt,'txt')
for x in file_list :
    file_txt_path = adr_txt + '\\' + x
    with open(file_txt_path,'r',encoding='utf8') as f:
        txt_firstline = f.readline()
        if txt_firstline.find('.') == -1  :
            erosion_dilation(x,adr_jpg_crop,adr_ocr_error)

file_list_ext(adr_ocr_error,'jpg')
for x in file_list :
    ocr(x,adr_ocr_error,adr_txt)

file_list_ext(adr_txt,'txt')
wb = openpyxl.load_workbook('arco_{}.xlsx'.format(today_date))
sheet = wb.active
for x in file_list:
    file_txt_path = adr_txt + '\\' + x
    a = a + 1
    with open(file_txt_path, 'r', encoding='utf8') as f:
        line = f.readline()
        if not line :
            line = 'none'
        sheet.cell(row=a, column=3).value = x[:-4]
        sheet.cell(row=a, column=4).value = line

wb.save('arco_{}.xlsx'.format(today_date))

print('finish')
